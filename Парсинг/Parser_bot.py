from selenium import webdriver
from bs4 import BeautifulSoup
import telebot
import openpyxl


option = webdriver.FirefoxOptions()
option.set_preference('general.useragent.override','Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0')

# --- Тут мы выключаем, что это Селениум веб драйвер
option.set_preference('dom.webdriver.enabled', False)



# ---- Данная опция позволяет работать Селениуму, не запуская веб морду.
option.headless = True

browser = webdriver.Firefox(options=option)
bot = telebot.TeleBot("1785979078:AAECq_sMxMZy5ANm0rPl5BAKUtEgWUqqQhg")

@bot.message_handler(commands=['start'])
def handle_start(message):
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        user_markup.row('/start','/stop')
        user_markup.row('Iphone','Samsung','Xiaomi')
        bot.send_message(message.from_user.id, 'Вас приветствует парсер бот kaspi, выберите какой бренд вас интересует',reply_markup = user_markup)

@bot.message_handler(commands=['stop'])
def handle_stop(message):
        hide_markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.from_user.id, '...', reply_markup=hide_markup)

@bot.message_handler(content_types=['text'])
def txt(message):
    if message.text == 'Samsung':
        bot.send_message(message.chat.id, 'Собираю информацию, ожидайте ответа ')
        count = 2
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(row=1, column=1).value = 'Смартфон'
        sheet.cell(row=1, column=2).value = 'Цена'
        try:
            for j in range(13):
                browser.get('https://kaspi.kz/shop/c/smartphones/brand-samsung/?at=1&page=' + str(j))
                # browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]').find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]/div[1]/div/ul[1]/li[8]/a').click()
                # ------ тут мы вытаскием уже содержимое тело сайта из бразуера, для парсинга в bs4
                page = browser.page_source

                soup = BeautifulSoup(page, 'lxml')
                s0 = soup.find('div', class_='item-cards-grid__cards').find_all('div', class_="item-card__info")
                # strip('\n') - убирает в тексте \n
                for x in s0:
                    sheet.cell(row=(count), column=1).value = x.find(class_="item-card__name").text.strip('\n')
                    sheet.cell(row=(count), column=2).value = x.find(class_="item-card__debet").find(class_="item-card__prices-price").text.strip('\n')
                    count += 1
            book.save('samsung.xlsx')
            book.close()
            document = open('samsung.xlsx', 'rb')
            bot.send_chat_action(message.from_user.id, 'upload_document')
            bot.send_document(message.from_user.id, document)

        except:
            bot.send_message(message.chat.id, 'Cтраница не найдена')

    if message.text == 'Iphone':
        bot.send_message(message.chat.id, 'Собираю информацию, ожидайте ответа ')
        count = 2
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(row=1, column=1).value = 'Смартфон'
        sheet.cell(row=1, column=2).value = 'Цена'
        try:
            for j in range(19):
                browser.get('https://kaspi.kz/shop/c/smartphones/class-apple/?page={}'.format(str(j)))
                # browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]').find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]/div[1]/div/ul[1]/li[8]/a').click()
                # ------ тут мы вытаскием уже содержимое тело сайта из бразуера, для парсинга в bs4
                page = browser.page_source

                soup = BeautifulSoup(page, 'lxml')
                s0 = soup.find('div', class_='item-cards-grid__cards').find_all('div', class_="item-card__info")
                # strip('\n') - убирает в тексте \n
                for x in s0:
                    sheet.cell(row=(count), column=1).value = x.find(class_="item-card__name").text.strip('\n')
                    sheet.cell(row=(count), column=2).value = x.find(class_="item-card__debet").find(class_="item-card__prices-price").text.strip('\n')
                    count += 1
            book.save('iphone.xlsx')
            book.close()
            document = open('iphone.xlsx', 'rb')
            bot.send_chat_action(message.from_user.id, 'upload_document')
            bot.send_document(message.from_user.id, document)

        except:
            bot.send_message(message.chat.id, 'Cтраница не найдена')
    if message.text == 'Xiaomi':
        bot.send_message(message.chat.id, 'Собираю информацию, ожидайте ответа ')
        count = 2
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(row=1, column=1).value = 'Смартфон'
        sheet.cell(row=1, column=2).value = 'Цена'
        try:
            for j in range(11):
                browser.get('https://kaspi.kz/shop/c/smartphones/brand-xiaomi/?q=%3Acategory%3ASmartphones&page={}'.format(str(j)))
                # browser.find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]').find_element_by_xpath('/html/body/div[6]/div[1]/div/div[1]/div[1]/div/ul[1]/li[8]/a').click()
                # ------ тут мы вытаскием уже содержимое тело сайта из бразуера, для парсинга в bs4
                page = browser.page_source

                soup = BeautifulSoup(page, 'lxml')
                s0 = soup.find('div', class_='item-cards-grid__cards').find_all('div', class_="item-card__info")
                # strip('\n') - убирает в тексте \n
                for x in s0:
                    sheet.cell(row=(count), column=1).value = x.find(class_="item-card__name").text.strip('\n')
                    sheet.cell(row=(count), column=2).value = x.find(class_="item-card__debet").find(class_="item-card__prices-price").text.strip('\n')
                    count += 1
            book.save('xiaomi.xlsx')
            book.close()
            document = open('xiaomi.xlsx', 'rb')
            bot.send_chat_action(message.from_user.id, 'upload_document')
            bot.send_document(message.from_user.id, document)

        except:
            bot.send_message(message.chat.id, 'Cтраница не найдена')
    else:
        bot.send_message(message.chat.id, 'Выберите из телефон из предложенных вариантов')
bot.polling(none_stop=True, interval = 0)
